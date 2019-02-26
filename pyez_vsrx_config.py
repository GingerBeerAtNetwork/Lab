#! /usr/bin/env python3

#导入相关模块

#excel处理模块
import openpyxl

#jinja2模块
from jinja2 import Environment, FileSystemLoader

#PyEZ模块
from jnpr.junos import Device
from jnpr.junos.utils.config import Config

#字典模块
from collections import defaultdict

#时间模块
from time import sleep

#XML处理模块
from lxml import etree

#定义一些空容器作为后续数据存储之用
config_container=[]
device_node_ip={}



#开始导入EXCEL文件，并根据工作表名称导入相关内容。
vsrx_config_excel= openpyxl.load_workbook("实验工作表.xlsx")

vsrx_int_config=vsrx_config_excel.get_sheet_by_name("接口配置")

vsrx_policy_config=vsrx_config_excel.get_sheet_by_name("安全策略")



#开始处理Jinja2，并导入相关参数。
jinja_loader = FileSystemLoader('./')

env = Environment(loader=jinja_loader)

vsrx_int_template = env.get_template('vsrx_int_config.j2')
vsrx_secpolicy_template = env.get_template('vsrx_security_policy.j2')


#通过工作表获取相关变量参数内容。
each=2

while each <= vsrx_int_config.max_row:
    device_name=vsrx_int_config["a"+str(each)].value
    device_ip=vsrx_int_config["b"+str(each)].value
    config_paramter={"int_name":vsrx_int_config["c"+str(each)].value,
                            "int_ipv4_addr":vsrx_int_config["d"+str(each)].value,
                            "area_number":vsrx_int_config["G"+str(each)].value,
                            "zone_name":vsrx_int_config["H"+ str(each)].value}
    routing_config=vsrx_int_template.render(config_paramter)
    config_container.append({device_name:routing_config})
    device_node_ip[device_name]=device_ip

    each+=1
each=2

while each <= vsrx_policy_config.max_row:
    device_name=vsrx_policy_config["a"+str(each)].value
    config_paramter={"source_zone":vsrx_policy_config["C"+str(each)].value,
                            "dest_zone":vsrx_policy_config["D"+str(each)].value,
                            "source_addr":vsrx_policy_config["F"+str(each)].value,
                            "dest_addr":vsrx_policy_config["G"+str(each)].value,
                            "app":vsrx_policy_config["H"+str(each)].value,
                            "action":vsrx_policy_config["I"+str(each)].value}
    secpolicy_config=vsrx_secpolicy_template.render(config_paramter)
    config_container.append({device_name:secpolicy_config})
    each+=1 

print("device_node_ip为：{}".format(device_node_ip))
print("#"*10)
print("config_container为：{}".format(config_container))
print("#"*10)



#获取设备列表和对应的IP地址，并去重。
final_config=defaultdict(list)

for each in config_container:
    for key,value in each.items():
        final_config[key].append(value)

#开始执行PyEZ连接和配置程序
for name,ip in device_node_ip.items():
    print("#"*10)
    print("#"*10)
    print("设备：{}， IP地址：{}".format(name,ip))
    print("#"*10)
    vSRX=Device(host=ip, user="gingerbeer", passwd="juniper123").open()

    vSRX_config=Config(vSRX, mode="exclusive") 
    print("#"*10)
    print("以下配置将被导入进设备{}：".format(name))
    print("#"*10)
    for each in final_config[name]:
        print(each)
        vSRX_config.load(each,format="set")
    if vSRX_config.diff() == None:
        print("无新增配置，连接关闭。")
        vSRX.close()
    else:
        print("{} show | compare输出内容如下，请确认：".format(name))
        vSRX_config.pdiff()
        yes_no=input("请输入yes 或者 no来确定是否提交上述配置：")
        while not(yes_no == "yes" or yes_no =="no"):
            print("输入错误，请重新输入yes或no。")
            yes_no=input("请输入yes或者no来确定是否提交上述配置：")
        if yes_no == "yes":

            vSRX_config.commit(comment="add ospf and security config by PyEZ",timeout=60)
            vSRX.close()

        else:
            print("连接关闭，谢谢!")
            vSRX.close()

#进入核查程序，核查OSPF邻居状态。            
print("")
print("配置部分结束，进入核查程序。")
print("")
print("程序暂停30秒，等待OSPF收敛完成。")
sleep(30)
print("")
print("开始执行OSPF邻居核查程序：")
for name,ip in device_node_ip.items():
    print("#"*10)
    print("#"*10)
    print("设备：{}， IP地址：{}".format(name,ip))
    vSRX=Device(host=ip, user="gingerbeer", passwd="juniper123").open()
    vSRX.rpc.get_ospf_neighbor_information()
    ospf_neighbor_status=etree.tostring(vSRX.rpc.get_ospf_neighbor_information()).decode("utf-8")
    print(ospf_neighbor_status)

print("程序执行完成。")

